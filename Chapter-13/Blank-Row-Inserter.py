#!python3
#blankRowInserter.py - takes two integers M and N and a filename as arguments and  
#insert M blank rows into the spreadsheet starting at row N

import openpyxl,sys
if len(sys.argv) > 1:
    N = int(sys.argv[1])
    M = int(sys.argv[2])
    filename = sys.argv[3]
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    tuple = tuple(sheet)
    count = 0
    for i in tuple:
        for j in i:
            c = j.column
            r = j.row
            value = sheet.cell(row=r, column=c).value 
            print(value)
            sheet.cell(row = r,column=c).value = " "
            if r < N:
                sheet2.cell(row=r, column=c).value = value
            if r>=N:
                sheet2.cell(row=r+M, column=c).value = value
            

    wb2.save(filename+'2.xlsx')
        

    
else:
    print("Enter arguement n!")