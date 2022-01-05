import os,sys
import openpyxl,sys

if len(sys.argv) > 1:
    path = sys.argv[1]
    row,column = 1,1;
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    tuple = tuple(sheet)
    for i in range(1,sheet.max_column+1):
        file = open('textfiles2/textfile{}.txt'.format(i), 'a')
        for j in range(1,sheet.max_row+1):
            value = sheet.cell(row = j , column= i).value
            file.write(str(value))
        file.close()

else:
    print("Enter arguement n!")
