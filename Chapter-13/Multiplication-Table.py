#! python3
# multiplicationTable.py - Tabulates multiplication table upto n numbers into a xlsx file.
# give n as arguement

import openpyxl,sys
from openpyxl.styles import Font
n=0;
if len(sys.argv) > 1:
    n = sys.argv[1]
    n = int(n)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Multiplication table upto "+str(n)
    for i in range(1,n+1):
        for j in range(1,n+1):
            if i==1 and j>1:
                sheet.cell(row=i, column=j).value = j-1
                sheet.cell(row=i, column=j).font = Font(bold=True)
            if j==1 and i>1:
                sheet.cell(row=i ,column=j).value = i-1
                sheet.cell(row=i ,column=j).font = Font(bold=True)
            sheet.cell(row=i+1, column=j+1).value = i*j
            sheet.cell(row=n+1,column=1).value=n
            sheet.cell(row=1,column=n+1).value=n
            sheet.cell(row=n+1,column=1).font=Font(bold=True)
            sheet.cell(row=1,column=n+1).font=Font(bold=True)
    wb.save('multiplicationTable.xlsx')
else:
    print("Enter arguement n!")